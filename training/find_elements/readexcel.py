from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment

wb = Workbook()
ws_ = wb.active
ws_.title = 'students_info'


ws_['A1'] = 'name'
ws_['B1'] = 'place'
ws_['C1'] = 'age'
ws_['D1'] = 'city'
ws_['A1'].font = Font(name = 'Calibri', bold= True, italic=True, color= 'FF0000')
# formatting entire row
bold_font = Font(bold=True, color='FF0000')
#fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
for cell in ws_[1]:
    cell.font = bold_font
    #cell.fill = fill
data_list = [['sweeta', 'Chennai', 25 , 'Tamilnadu'],
             ['Andrea', 'Hubbli', 21 , 'Karnataka'],
             ['Krish', 'Madhapur', 27, 'Hyderabad'],
             ['Radha', "barsana", 18, 'Uttarpradesh']]
for data in data_list:
    ws_.append(data)

font_style = Font(name= 'arial', color= '00FFFF', bold= False, italic= True)
for cell in ws_[3]:
    cell.font= font_style

wb.save('infoprac.xlsx')


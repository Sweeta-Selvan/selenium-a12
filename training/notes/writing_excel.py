'''
To write into the excel files, we use openpyxl

To install openpyxl
    Go to command prompt    -->     pip install openpyxl

'''

from openpyxl import Workbook

# ## create the excel file
# workbook = Workbook()
#
# ## initialize the active worksheet
# worksheet = workbook.active
#
# ## set the title for the worksheet(optional)
# worksheet.title = 'personal_info'
#
# ## Add the data to the excel
# worksheet['A1'] = 'name'
# worksheet['B1'] = 'place'
# worksheet['C1'] = 'email_id'
#
# data_list = [
#     ['Arsha', 'Bengaluru', 'arsha@gmail.com'],
#     ['Ankitha', 'Mysore', 'ankitha@gmail.com'],
#     ['Ashwini', 'Hubli', 'ashwini@gmail.com'],
#     ['Sujitha', 'Dharwad', 'sujitha@gmail.com'],
#     ['Pooja', 'Tumkur', 'pooja@gmail.com']
# ]
#
# for data in data_list:
#     worksheet.append(data)
#
# # ## save the file
# # workbook.save('A12_candidates.xlsx')
#
# ## save the file in some other location
# workbook.save(r'C:\Users\Ramya\PycharmProjects\Sel-A12-June16-2025\files\A12_candidates.xlsx')


####################################################################################

## ANALYZE THE BELOW CODE

from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment

## create the excel workbook
excel_workbook = Workbook()

## initialize the worksheet
worksheet = excel_workbook.active

## set a name for the worksheet(optional)
worksheet.title = 'personal_information'

## Adding the data
worksheet['A1'] = 'name'
worksheet['B1'] = 'place'
worksheet['C1'] = 'salary'
worksheet['D1'] = 'email_id'

# ## Font formatting
# worksheet['A1'].font = Font(name='Arial', bold=True, italic=True, color='FF0000')   ## formats individual cell

#
data_list = [
    ['Vihaan', 'Delhi', 50000, 'vihaan@gmail.com'],
    ['Siddharth', 'Mumbai', 55000, 'siddu@gmail.com'],
    ['Saina', 'Hyderabad', 60000, 'saina@gmail.com'],
    ['Risto', 'Pune', 65000, 'risto@gmail.com'],
    ['Goku', 'Assam', 70000, 'goku@gmail.com'],
    ['Paresh', 'Gujrat', 75000, 'paresh@gmail.com']
]

for row in data_list:
    worksheet.append(row)


# ##-----------------------------------------------------------------------------
# ## formatting entire row
# ## define styles
# bold_font = Font(bold=True, color='FFFFFF')
# fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
# for cell in worksheet[1]:
#     cell.font = bold_font
#     cell.fill = fill
#
# excel_workbook.save('example.xlsx')

# ##-----------------------------------------------------------------------------

# ## Add new sheet to the existing workbook
#
# ## load existing workbook
# workbook = load_workbook('emp_data.xlsx')
#
# ## create a new sheet
# new_sheet = workbook.create_sheet(title="title")
#
# ## write the data asusal
# ## save the file
#
# ##-----------------------------------------------------------------------------
#
# ## save the excel file
# excel_workbook.save('emp_data.xlsx')
#
# # ## save the excel file in different location
# # excel_workbook.save(r'C:\Users\Ramya\PycharmProjects\Sel-M6-April1\files_\emp_data.xlsx')
#
# ##-----------------------------------------------------------------------------
#
# ## Clear the cell's value
#
# from openpyxl import load_workbook
#
# workbook = load_workbook("emp_data.xlsx")
# worksheet = workbook.active
#
# worksheet['A1'] = None      ## Clears the value and formats in cell A1
#
# workbook.save("emp_data.xlsx")
#
# ##-----------------------------------------------------------------------------
#
# ## Delete entire row or column
#
# from openpyxl import load_workbook
#
# workbook = load_workbook("emp_data.xlsx")
# worksheet = workbook.active
#
# worksheet.delete_rows(1)        ## Deletes row1
# worksheet.delete_cols(1)        ## Deletes col1



































